"""
Export a pipeline run to a formatted Excel workbook using openpyxl.
"""
import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

FIELDS = [
    "loan_amount", "beneficiary_name", "account_number", "bank_name",
    "document_date", "reference_number", "interest_rate", "loan_type",
    "maturity_date", "credit_limit", "currency", "purpose",
    "guarantor_name", "collateral",
]
FIELD_LABELS = {
    "loan_amount": "Loan Amount",
    "beneficiary_name": "Beneficiary Name",
    "account_number": "Account Number",
    "bank_name": "Bank Name",
    "document_date": "Document Date",
    "reference_number": "Reference Number",
    "interest_rate": "Interest Rate",
    "loan_type": "Loan Type",
    "maturity_date": "Maturity Date",
    "credit_limit": "Credit Limit",
    "currency": "Currency",
    "purpose": "Purpose",
    "guarantor_name": "Guarantor Name",
    "collateral": "Collateral",
}

# ── Styles ────────────────────────────────────────────────────────────────────
def _hdr_font(): return Font(bold=True, color="FFFFFF", size=11)
def _hdr_fill(hex_color="1E293B"): return PatternFill("solid", fgColor=hex_color)
def _title_font(): return Font(bold=True, size=14, color="1E293B")
def _label_font(): return Font(bold=True, size=10, color="475569")
def _thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

STATUS_FILLS = {
    "correct":    PatternFill("solid", fgColor="D1FAE5"),
    "incorrect":  PatternFill("solid", fgColor="FEE2E2"),
    "missing":    PatternFill("solid", fgColor="FEF3C7"),
    "partial":    PatternFill("solid", fgColor="FFEDD5"),
    "acceptable": PatternFill("solid", fgColor="F1F5F9"),
}

def _accuracy_fill(acc: Optional[float]):
    if acc is None: return PatternFill("solid", fgColor="F1F5F9")
    if acc >= 85:   return PatternFill("solid", fgColor="D1FAE5")
    if acc >= 65:   return PatternFill("solid", fgColor="FEF3C7")
    return             PatternFill("solid", fgColor="FEE2E2")


def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_header_row(ws, row: int, headers: list[str], fill_color="1E293B"):
    fill = _hdr_fill(fill_color)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _hdr_font()
        cell.fill = fill
        cell.alignment = _center()
        cell.border = _thin_border()


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
def _sheet_summary(wb, run: dict):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Banking PDF Extractor — Pipeline Run Report"
    t.font = _title_font()
    t.alignment = _center()
    t.fill = PatternFill("solid", fgColor="EEF2FF")

    ws.merge_cells("A2:D2")

    pairs = [
        ("Run ID", run["run_id"]),
        ("Date", run["created_at"][:19].replace("T", " ")),
        ("Folder / Source", run["folder_path"]),
        ("Status", run["status"].upper()),
        ("Total Files", run["total_files"]),
        ("Succeeded", run["succeeded"]),
        ("Failed", run["failed"]),
        ("Average Accuracy", f"{run['avg_accuracy'] or 0:.1f}%"),
        ("Auto-tune", "Yes" if run["settings"].get("auto_tune", True) else "No"),
        ("Accuracy Target", f"{run['settings'].get('accuracy_threshold', 85)}%"),
        ("Max Iterations", run["settings"].get("max_iterations", 3)),
    ]
    for i, (label, value) in enumerate(pairs, start=3):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _label_font()
        lc.fill = PatternFill("solid", fgColor="F8FAFC")
        lc.alignment = _left()
        lc.border = _thin_border()

        vc = ws.cell(row=i, column=2, value=value)
        vc.alignment = _left()
        vc.border = _thin_border()
        if label == "Status":
            vc.fill = PatternFill("solid", fgColor="D1FAE5" if value == "COMPLETE" else "FEE2E2")
        if label == "Average Accuracy":
            vc.fill = _accuracy_fill(run["avg_accuracy"])

    _set_col_widths(ws, {"A": 22, "B": 38})


# ── Sheet 2: Extractions ──────────────────────────────────────────────────────
def _sheet_extractions(wb, run: dict):
    ws = wb.create_sheet("Extractions")
    ws.sheet_view.showGridLines = False

    headers = ["Filename", "Status", "Accuracy", "Initial Accuracy"] + \
              [FIELD_LABELS[f] for f in FIELDS] + ["Error"]
    _write_header_row(ws, 1, headers)
    ws.freeze_panes = "A2"

    for r, f in enumerate(run["files"], start=2):
        acc = f.get("accuracy")
        extraction = f.get("extraction") or {}
        cols = [
            f["filename"],
            f["status"].upper(),
            f"{acc:.1f}%" if acc is not None else "—",
            f"{f['initial_accuracy']:.1f}%" if f.get("initial_accuracy") is not None else "—",
        ] + [extraction.get(field) or "" for field in FIELDS] + [f.get("error") or ""]

        acc_fill = _accuracy_fill(acc)
        for c, val in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = _left()
            cell.border = _thin_border()
            if c in (1, 2, 3, 4):
                cell.fill = acc_fill

    # Column widths
    widths = {"A": 30, "B": 12, "C": 12, "D": 16}
    for i, f in enumerate(FIELDS, start=5):
        widths[get_column_letter(i)] = 22
    widths[get_column_letter(len(FIELDS) + 5)] = 40
    _set_col_widths(ws, widths)


# ── Sheet 3: Field Accuracy ───────────────────────────────────────────────────
def _sheet_field_accuracy(wb, run: dict):
    ws = wb.create_sheet("Field Accuracy")
    ws.sheet_view.showGridLines = False

    headers = ["Filename", "Overall Accuracy"] + [FIELD_LABELS[f] for f in FIELDS]
    _write_header_row(ws, 1, headers, fill_color="1E3A5F")
    ws.freeze_panes = "A2"

    for r, f in enumerate(run["files"], start=2):
        acc = f.get("accuracy")
        judge = f.get("judge_result") or {}
        field_scores = judge.get("field_scores") or {}

        ws.cell(row=r, column=1, value=f["filename"]).border = _thin_border()

        acc_cell = ws.cell(row=r, column=2,
                           value=f"{acc:.1f}%" if acc is not None else "—")
        acc_cell.fill = _accuracy_fill(acc)
        acc_cell.border = _thin_border()
        acc_cell.alignment = _center()

        for c, field in enumerate(FIELDS, start=3):
            score = field_scores.get(field, {})
            status = score.get("status", "—")
            correct_val = score.get("correct_value") or ""
            cell = ws.cell(row=r, column=c,
                           value=f"{status}" + (f"\n→ {correct_val}" if correct_val and status != "correct" else ""))
            cell.fill = STATUS_FILLS.get(status, PatternFill("solid", fgColor="F1F5F9"))
            cell.border = _thin_border()
            cell.alignment = _left()

    widths = {"A": 30, "B": 18}
    for i in range(len(FIELDS)):
        widths[get_column_letter(i + 3)] = 20
    _set_col_widths(ws, widths)

    # Legend
    legend_row = len(run["files"]) + 3
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True, size=9)
    for i, (status, fill) in enumerate(STATUS_FILLS.items(), start=1):
        c = ws.cell(row=legend_row + i, column=1, value=f"  {status.capitalize()}")
        c.fill = fill
        c.font = Font(size=9)
        c.border = _thin_border()


# ── Sheet 4: Autotune Trajectory ──────────────────────────────────────────────
def _sheet_autotune(wb, run: dict):
    ws = wb.create_sheet("Autotune Trajectory")
    ws.sheet_view.showGridLines = False
    headers = ["Filename", "Iteration", "Accuracy", "Feedback"]
    _write_header_row(ws, 1, headers, fill_color="065F46")
    ws.freeze_panes = "A2"

    r = 2
    for f in run["files"]:
        iters = f.get("autotune_iters") or []
        for it in iters:
            acc = it.get("accuracy")
            judge = it.get("judge_result") or {}
            ws.cell(row=r, column=1, value=f["filename"]).border = _thin_border()
            ws.cell(row=r, column=2, value=it.get("iteration")).border = _thin_border()
            acc_c = ws.cell(row=r, column=3,
                            value=f"{acc:.1f}%" if acc is not None else "—")
            acc_c.fill = _accuracy_fill(acc)
            acc_c.border = _thin_border()
            ws.cell(row=r, column=4, value=judge.get("overall_feedback", "")).border = _thin_border()
            r += 1

    _set_col_widths(ws, {"A": 30, "B": 12, "C": 14, "D": 80})


# ── Public API ────────────────────────────────────────────────────────────────
def build_excel(run: dict) -> bytes:
    wb = openpyxl.Workbook()
    _sheet_summary(wb, run)
    _sheet_extractions(wb, run)
    _sheet_field_accuracy(wb, run)
    _sheet_autotune(wb, run)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
